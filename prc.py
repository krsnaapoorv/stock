import openpyxl
import pandas as pd
import xlsxwriter
import datetime
import copy
path = "/home/apoorva/QED/Nifty 100.xlsx"
s = []


# opening Sheet of xlsx file
wb_obj = openpyxl.load_workbook(path) 
sheet_obj = wb_obj.active 


row = sheet_obj.max_row
total_return = []

# Finding all the listed company and their accord no.
for i in range(4,row,315):
    accord_no = sheet_obj.cell(row = i, column = 2)
    company_name = sheet_obj.cell(row = i, column = 3)
    s.append({company_name.value: accord_no.value})


j,sum = 0,0
 

#  Finding daily return value
def six_month_return(start_month,no_of_working_day):
    k,sum  = 0,0
    total_return = []
    for i in range(4,row):
        company_name = sheet_obj.cell(row = i, column = 3).value
        company_name_next_row = sheet_obj.cell(row = i, column = 3).value
        each_date = sheet_obj.cell(row = i, column = 4).value
        next_m = sheet_obj.cell(row = i, column = 4).value
        if each_date != None and next_m != None and company_name == company_name_next_row:
            nifty_year = each_date.year
            nifty_month = each_date.month
            next_month = next_m.month
            month_range = []
            for m in range(start_month,start_month+6):
                month_range.append(m)
            if nifty_year == 2019 and nifty_month in month_range and next_month in month_range and next_month == next_month:
                k += 1
                closing_value_present_day = sheet_obj.cell(row = i, column = 8).value
                closing_value_last_day = sheet_obj.cell(row = i+1, column = 8).value
                if(closing_value_present_day !=  None and closing_value_last_day != None):
                    return_value_each_day = (closing_value_present_day - closing_value_last_day)
                    # if i < 299:
                    #     print(i)
                    sum += return_value_each_day
                    if k == no_of_working_day:
                        total_return.append(sum/k)
                        # print(company_name,sum,k)
            else:
                sum = 0
                k = 0
    return total_return


def six_month_return_mixed_year(start_month,no_of_working_day):
    k,sum  = 0,0
    z = 0
    total_return = []
    for i in range(4,row):
        company_name = sheet_obj.cell(row = i, column = 3).value
        company_name_next_row = sheet_obj.cell(row = i, column = 3).value
        each_date = sheet_obj.cell(row = i, column = 4).value
        next_m = sheet_obj.cell(row = i, column = 4).value
        if each_date != None and next_m != None and company_name == company_name_next_row:
            nifty_year = each_date.year
            nifty_month = each_date.month
            next_month = next_m.month
            month_range = []
            month_range_2020 = []
            for m in range(start_month,start_month+6):
                if m <= 12:
                    month_range.append(m)
                else:
                    month_range_2020.append(m-12)
            if nifty_year == 2019 and nifty_month in month_range and next_month in (month_range+month_range_2020) and next_month == next_month:
                k += 1
                closing_value_present_day = sheet_obj.cell(row = i, column = 8).value
                closing_value_last_day = sheet_obj.cell(row = i+1, column = 8).value
                if(closing_value_present_day !=  None and closing_value_last_day != None):
                    return_value_each_day = (closing_value_present_day - closing_value_last_day)
                    # if i < 299:
                    #     print(company_name)
                    # # print(company_name)
                    sum += return_value_each_day
                    if k == no_of_working_day or (company_name == "Punjab National Bank" and k == no_of_working_day-1):
                        # z += 1
                        total_return.append(sum/k)
                        # print(company_name,sum,k,z)
            elif nifty_year == 2020 and nifty_month in month_range_2020 and next_month in (month_range+month_range_2020) and next_month == next_month:
                k += 1
                closing_value_present_day = sheet_obj.cell(row = i, column = 8).value
                closing_value_last_day = sheet_obj.cell(row = i+1, column = 8).value
                if(closing_value_present_day !=  None and closing_value_last_day != None):
                    return_value_each_day = (closing_value_present_day - closing_value_last_day)
                    # if i < 299:
                    #     print(company_name)
                    # print(company_name)
                    sum += return_value_each_day
                    if k == no_of_working_day or (company_name == "Punjab National Bank" and k == no_of_working_day-1):
                        # z += 1
                        total_return.append(sum/k)
                        # print(company_name,sum,k,z)
            else:
                sum = 0
                k = 0
    return total_return


def Momentum_rank(arr,initial_arr):
    distinct_s = copy.deepcopy(s)
    for (i,j) in zip(distinct_s,arr):
        i["average_return"] = j
    distinct_s.pop()
    li = sorted(distinct_s, key = lambda i: i["average_return"],reverse=True)
    k = 0
    for i in li:
        k += 1
        i["Momentum_rank"] = k
    # print(li)
    return li


li_of_no_of_working_day_for_every_six_month_2019 = [121,121,121,121,122,120,102,122,122,125]

total_return_june = six_month_return(1,121)
momentum_june = Momentum_rank(total_return_june,s)

total_return_july = six_month_return(2,121)
momentum_july = Momentum_rank(total_return_july,s)

total_return_aug = six_month_return(3,121)
momentum_aug = Momentum_rank(total_return_aug,s)

total_return_sep = six_month_return(4,121)
momentum_sep = Momentum_rank(total_return_sep,s)

total_return_oct = six_month_return(5,122)
momentum_oct = Momentum_rank(total_return_oct,s)

total_return_nov = six_month_return(6,120)
momentum_nov = Momentum_rank(total_return_nov,s)

total_return_dec = six_month_return(7,102)
momentum_dec = Momentum_rank(total_return_dec,s)

total_return_jan = six_month_return_mixed_year(8,122)
momentum_jan = Momentum_rank(total_return_jan,s)

total_return_feb = six_month_return_mixed_year(9,122)
momentum_feb = Momentum_rank(total_return_feb,s)

total_return_march = six_month_return_mixed_year(10,125)
print(len(total_return_march))
momentum_march = Momentum_rank(total_return_march,s)


# Seperating each column from the list of company dictionary data
def listing(arr):
    companies = []
    accord_no = []
    average_return = []
    momentum_rank = []
    for i in arr:
        company = list(i.keys())[0]
        companies.append(company)
        accord_no.append(i[company])
        average_return.append(i["average_return"])
        momentum_rank.append(i["Momentum_rank"])
    return [companies,accord_no,average_return,momentum_rank]

arr_june = listing(momentum_june)
arr_july = listing(momentum_july)
arr_aug = listing(momentum_aug)
arr_sep = listing(momentum_sep)
arr_oct = listing(momentum_oct)
arr_nov = listing(momentum_nov)
arr_dec = listing(momentum_dec)
arr_jan = listing(momentum_jan)
arr_feb = listing(momentum_feb)
arr_march = listing(momentum_march)

# Writing dataframe to be written in xlsx shet
df_june = pd.DataFrame({'Company Name': arr_june[0],"Acoord_No": arr_june[1],"Average_return": arr_june[2], "Momentum Rank": arr_june[3]})
df_july = pd.DataFrame({'Company Name': arr_july[0],"Acoord_No": arr_july[1],"Average_return": arr_july[2], "Momentum Rank": arr_july[3]})
df_aug = pd.DataFrame({'Company Name': arr_aug[0],"Acoord_No": arr_aug[1],"Average_return": arr_aug[2], "Momentum Rank": arr_aug[3]})
df_sep = pd.DataFrame({'Company Name': arr_sep[0],"Acoord_No": arr_sep[1],"Average_return": arr_sep[2], "Momentum Rank": arr_sep[3]})
df_oct = pd.DataFrame({'Company Name': arr_oct[0],"Acoord_No": arr_oct[1],"Average_return": arr_oct[2], "Momentum Rank": arr_oct[3]})
df_nov = pd.DataFrame({'Company Name': arr_nov[0],"Acoord_No": arr_nov[1],"Average_return": arr_nov[2], "Momentum Rank": arr_nov[3]})
df_dec = pd.DataFrame({'Company Name': arr_dec[0],"Acoord_No": arr_dec[1],"Average_return": arr_dec[2], "Momentum Rank": arr_dec[3]})
df_jan = pd.DataFrame({'Company Name': arr_jan[0],"Acoord_No": arr_jan[1],"Average_return": arr_jan[2], "Momentum Rank": arr_jan[3]})
df_feb = pd.DataFrame({'Company Name': arr_feb[0],"Acoord_No": arr_feb[1],"Average_return": arr_feb[2], "Momentum Rank": arr_feb[3]})
df_march = pd.DataFrame({'Company Name': arr_march[0],"Acoord_No": arr_march[1],"Average_return": arr_march[2], "Momentum Rank": arr_march[3]})


writer = pd.ExcelWriter('every_month.xlsx', engine='xlsxwriter')

# Converting the dataframe to an XlsxWriter Excel object.
df_june.to_excel(writer, sheet_name='June')
df_july.to_excel(writer, sheet_name='July')
df_aug.to_excel(writer, sheet_name='aug')
df_sep.to_excel(writer, sheet_name='sep')
df_oct.to_excel(writer, sheet_name='oct')
df_nov.to_excel(writer, sheet_name='nov')
df_dec.to_excel(writer, sheet_name='dec')
df_jan.to_excel(writer, sheet_name='jan')
df_feb.to_excel(writer, sheet_name='feb')
df_march.to_excel(writer, sheet_name='march')

# Close the Pandas Excel writer and output the Excel file.
writer.save()

