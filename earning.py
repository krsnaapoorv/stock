import openpyxl
import matplotlib.pyplot as plt

path = "/home/apoorva/QED/every_month.xlsx"


wb_obj = openpyxl.load_workbook(path) 

# Finding the gain loss on the basis of momentum rank sheet by comparing with previous value
def my_return_every_month(last_month,present_month,investment):

    per_company_investment = investment/20
    sheet_obj = wb_obj[last_month] 
    sheet_obj_next = wb_obj[present_month]

    row = sheet_obj.max_row

    s = []
    s_next = []
    for i in range(1,row):
        accord_no = sheet_obj.cell(row = i, column = 2)
        accord_no_next = sheet_obj_next.cell(row = i, column = 2)
        average_return = sheet_obj.cell(row = i, column = 3)
        average_return_next = sheet_obj_next.cell(row = i, column = 3)
        company_name = sheet_obj.cell(row = i, column = 4)
        company_name_next = sheet_obj_next.cell(row = i, column = 4)
        Momentum_rank = sheet_obj.cell(row = i, column = 5)
        Momentum_rank_next = sheet_obj_next.cell(row = i, column = 5)
        s.append({"Company Name":company_name.value,"Accord_value": accord_no.value,"average_return":average_return.value,"Momentum_rank":Momentum_rank.value})
        s_next.append({"Company Name":company_name_next.value,"Accord_value": accord_no_next.value,"average_return":average_return_next.value,"Momentum_rank":Momentum_rank_next.value})

    # print(s)
    present_month_return_top_20 = []
    next_month_return_top_20 = []
    for i in range(21):
        present_month_return_top_20.append(s[i])
        next_month_return_top_20.append(s_next[i])


    amount_change_after_one_month = 0

    for i in range(1,21):
        for j in range(1,row-1):
            if s[i]["Accord_value"] == s_next[j]["Accord_value"]:
                increment = s_next[j]["average_return"] - s[i]["average_return"]
                # print(increment/100,i,j)
                # print(increment)
                amount_change_after_one_month +=  (per_company_investment*(increment/100))

    remaining_amount = investment + amount_change_after_one_month
    return [remaining_amount,amount_change_after_one_month]

gain_loss = []
x = []
y = []

july = my_return_every_month("June","July",1000000)
gain_loss.append(["july",july[1]])
y.append(july[0])
aug = my_return_every_month("July","aug",july[0])
gain_loss.append(["aug",aug[1]])
y.append(aug[0])
sep = my_return_every_month("aug","sep",aug[0])
gain_loss.append(["sep",sep[1]])
y.append(sep[0])
octo = my_return_every_month("sep","oct",sep[0])
gain_loss.append(["octo",octo[1]])
y.append(octo[0])
nov = my_return_every_month("oct","nov",octo[0])
gain_loss.append(["nov",nov[1]])
y.append(nov[0])
dec = my_return_every_month("nov","dec",nov[0])
gain_loss.append(["dec",dec[1]])
y.append(dec[0])
jan = my_return_every_month("dec","jan",dec[0])
gain_loss.append(["jan",jan[1]])
y.append(jan[0])
feb = my_return_every_month("jan","feb",jan[0])
gain_loss.append(["feb",feb[1]])
y.append(feb[0])
march = my_return_every_month("feb","march",feb[0])
gain_loss.append(["march",march[1]])
y.append(march[0])
# print(march[0])
# print(gain_loss)


# Plotting Monthly return


for i in gain_loss:
    x.append(i[0])



plt.plot(x, y) 
plt.xlabel('month')
plt.ylabel('amount')

plt.title('Equity curve')

plt.show()
